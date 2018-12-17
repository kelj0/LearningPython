#!/usr/bin/python3

# sendDuesReminders.py - sends emails based on payment status in spreadsheet
# Usage-> When Running script first arg is email second is password :)

import openpyxl,smtplib,sys

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.get_highes_column()
lastestMonth = sheet.cell(row=1,column=lastCol).value

unpaidMembers = {}
for r in range(2,sheet.get_highest_row()+1):
    payment = sheet.cell(row=r,column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r,column=1).value
        email=sheet.cell(row=r,column=2).value
        unpaidMembers[name]=email
smtpObj = smtplib.SMTP('smtp.gmail.com',587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login(sys.argv[1],sys.argv[2])

for name,email in unpaidMembers.items():
    body="Subject: %s dues unpaid.\nDeanDear %s,\nRecords show that you have not\
    paid dues for %s. Please make this payment as soon as possible. Thank you!'"%(lastestMonth, name, lastestMonth)
    print('Sending email to %s...'% email)
    sendmailStatus = smtpObj.sendmail(sys.argv[1],email,body)

    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s'%(email,sendmailStatus))

smtpObj.quit()
